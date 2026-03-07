#!/usr/bin/env python3
"""
GHG Data Validation and Canonical Dataset Generator

This script:
1. Parses the Excel source file as ground truth
2. Generates clean canonical CSV for the dashboard
3. Validates dashboard values against Excel
4. Reports any mismatches

Ground truth values from Excel:
- 2567 total GHG = 220.98693744 tCO2e
- 2568 total GHG = 231.620303712 tCO2e
- YoY total = +10.633366272 tCO2e = +4.81%
"""

import openpyxl
from openpyxl import load_workbook
import csv
import json
import os
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict

# Ground truth from Excel
GROUND_TRUTH = {
    '2567_total_tco2e': Decimal('220.98693744'),
    '2568_total_tco2e': Decimal('231.620303712'),
    'yoy_diff_tco2e': Decimal('10.633366272'),
    'yoy_pct': Decimal('4.81'),
    '2568_scopes': {
        'Scope1': Decimal('10.847924292'),
        'Scope2': Decimal('201.47809632'),
        'Scope3': Decimal('19.29428310'),
    },
    '2567_scopes': {
        'Scope1': Decimal('11.01713228'),
        'Scope2': Decimal('192.53468536'),
        'Scope3': Decimal('17.43511980'),
    },
    '2568_top_activities': [
        ('electricity', Decimal('201.47809632')),
        ('landfill waste', Decimal('10.16392')),
        ('septic tank CH4', Decimal('7.8204')),
        ('paper', Decimal('4.6197756')),
        ('provincial water', Decimal('4.5105875')),
    ],
    '2568_monthly_kgco2e': {
        1: Decimal('11530.49'),
        2: Decimal('14451.72'),
        3: Decimal('21384.86'),
        4: Decimal('20864.64'),
        5: Decimal('22672.85'),
        6: Decimal('21784.21'),
        7: Decimal('21964.51'),
        8: Decimal('22233.70'),
        9: Decimal('23440.91'),
        10: Decimal('18851.13'),
        11: Decimal('18425.95'),
        12: Decimal('14015.33'),
    }
}

def normalize_scope(scope_str):
    """Normalize scope string to standard format."""
    if not scope_str:
        return None
    s = str(scope_str).strip().lower()
    if 'scope 1' in s or 'scope1' in s:
        return 'Scope 1 (ประเภท 1)'
    elif 'scope 2' in s or 'scope2' in s:
        return 'Scope 2 (ประเภท 2)'
    elif 'scope 3' in s or 'scope3' in s:
        return 'Scope 3 (ประเภท 3)'
    return scope_str.strip()

def parse_excel(excel_path):
    """Parse Excel file and extract clean GHG data."""
    print(f"Reading Excel: {excel_path}")
    
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    data_rows = []
    header = None
    
    # Find the actual header row (look for 'year' or 'ปี' in first few rows)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        row_values = [str(c).strip() if c else '' for c in row]
        if any('year' in v.lower() or 'ปี' in v.lower() for v in row_values if v):
            header = row_values
            header_row = row_idx
            print(f"Found header at row {header_row}: {header[:5]}...")
            break
    
    if not header:
        print("ERROR: Could not find header row")
        return [], []
    
    # Extract data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        row_values = [str(c).strip() if c is not None else '' for c in row]
        
        # Skip empty rows or summary rows
        if not row_values[0]:
            continue
        
        # Skip if first column is not a year (2567 or 2568)
        year_val = row_values[0]
        try:
            year = int(float(year_val))
            if year not in (2567, 2568):
                continue
        except (ValueError, TypeError):
            continue
        
        data_rows.append(row_values)
    
    return header, data_rows

def analyze_current_csv(csv_path):
    """Analyze the current CSV to find issues."""
    print(f"\nAnalyzing current CSV: {csv_path}")
    
    issues = []
    valid_rows = []
    
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        content = f.read()
    
    # Parse CSV
    lines = content.strip().split('\n')
    header = lines[0].split(',')
    
    print(f"Header: {header}")
    print(f"Total lines: {len(lines)}")
    
    for line_idx, line in enumerate(lines[1:], 2):
        # Skip empty lines
        if not line.strip():
            continue
        
        # Count commas to check column count
        comma_count = line.count(',')
        expected_commas = len(header) - 1
        
        # Parse with proper CSV handling
        try:
            reader = csv.reader([line])
            row = next(reader)
            
            # Skip header rows (rows that contain 'year' in first column but not numeric)
            first_col = row[0].strip() if row else ''
            if first_col.lower() == 'year' or 'รายการ' in first_col:
                issues.append(f"Line {line_idx}: Embedded header row - '{first_col[:50]}'")
                continue
            
            # Skip summary rows (rows without valid scope)
            if len(row) > 2:
                scope = row[2].strip() if len(row) > 2 else ''
                activity = row[3].strip() if len(row) > 3 else ''
                
                # Check if this is a valid data row
                try:
                    year = int(float(row[0]))
                    month = int(float(row[1])) if len(row) > 1 and row[1] else 0
                    
                    if year in (2567, 2568) and month >= 1 and month <= 12:
                        valid_rows.append(row)
                    else:
                        issues.append(f"Line {line_idx}: Invalid year/month - year={year}, month={month}")
                except (ValueError, TypeError, IndexError) as e:
                    issues.append(f"Line {line_idx}: Parse error - {e}")
        except Exception as e:
            issues.append(f"Line {line_idx}: CSV parse error - {e}")
    
    print(f"\nValid rows: {len(valid_rows)}")
    print(f"Issues found: {len(issues)}")
    
    if issues[:10]:
        print("\nFirst 10 issues:")
        for issue in issues[:10]:
            print(f"  {issue}")
    
    return valid_rows, issues

def compute_totals_from_rows(rows):
    """Compute totals from valid rows."""
    year_totals = defaultdict(Decimal)
    year_month_totals = defaultdict(Decimal)
    year_scope_totals = defaultdict(lambda: defaultdict(Decimal))
    year_activity_totals = defaultdict(lambda: defaultdict(Decimal))
    
    for row in rows:
        try:
            year = int(float(row[0]))
            month = int(float(row[1]))
            scope = row[2].strip() if len(row) > 2 else ''
            
            # Find emission_tco2e value
            # Based on CSV structure, emission_tco2e is the last column
            emission_str = row[-1].strip() if row[-1] else '0'
            
            try:
                emission = Decimal(emission_str)
            except:
                continue
            
            year_totals[year] += emission
            year_month_totals[(year, month)] += emission
            
            # Normalize scope
            if 'scope 1' in scope.lower():
                year_scope_totals[year]['Scope1'] += emission
            elif 'scope 2' in scope.lower():
                year_scope_totals[year]['Scope2'] += emission
            elif 'scope 3' in scope.lower():
                year_scope_totals[year]['Scope3'] += emission
            
            # Activity
            activity = row[3].strip() if len(row) > 3 else ''
            if activity and activity not in ['รายการ', '']:
                year_activity_totals[year][activity] += emission
                
        except (ValueError, TypeError, IndexError) as e:
            continue
    
    return {
        'year_totals': dict(year_totals),
        'year_month_totals': {k: float(v) for k, v in year_month_totals.items()},
        'year_scope_totals': {y: dict(s) for y, s in year_scope_totals.items()},
        'year_activity_totals': {y: dict(a) for y, a in year_activity_totals.items()},
    }

def validate_against_ground_truth(computed):
    """Validate computed values against ground truth."""
    print("\n" + "="*60)
    print("VALIDATION AGAINST GROUND TRUTH")
    print("="*60)
    
    mismatches = []
    
    # 2567 total
    computed_2567 = Decimal(str(computed['year_totals'].get(2567, 0)))
    expected_2567 = GROUND_TRUTH['2567_total_tco2e']
    diff_2567 = abs(computed_2567 - expected_2567)
    match_2567 = diff_2567 < Decimal('0.01')
    
    print(f"\n2567 Total GHG:")
    print(f"  Computed: {computed_2567:.8f} tCO2e")
    print(f"  Expected: {expected_2567:.8f} tCO2e")
    print(f"  Diff:     {diff_2567:.8f} tCO2e")
    print(f"  Status:   {'PASS' if match_2567 else 'FAIL'}")
    
    if not match_2567:
        mismatches.append(f"2567 total mismatch: {computed_2567} vs {expected_2567}")
    
    # 2568 total
    computed_2568 = Decimal(str(computed['year_totals'].get(2568, 0)))
    expected_2568 = GROUND_TRUTH['2568_total_tco2e']
    diff_2568 = abs(computed_2568 - expected_2568)
    match_2568 = diff_2568 < Decimal('0.01')
    
    print(f"\n2568 Total GHG:")
    print(f"  Computed: {computed_2568:.8f} tCO2e")
    print(f"  Expected: {expected_2568:.8f} tCO2e")
    print(f"  Diff:     {diff_2568:.8f} tCO2e")
    print(f"  Status:   {'PASS' if match_2568 else 'FAIL'}")
    
    if not match_2568:
        mismatches.append(f"2568 total mismatch: {computed_2568} vs {expected_2568}")
    
    # YoY
    if computed_2567 > 0:
        computed_yoy = float((computed_2568 - computed_2567) / computed_2567 * 100)
        expected_yoy = float(GROUND_TRUTH['yoy_pct'])
        diff_yoy = abs(computed_yoy - expected_yoy)
        match_yoy = diff_yoy < 0.1
        
        print(f"\nYoY %:")
        print(f"  Computed: {computed_yoy:.2f}%")
        print(f"  Expected: {expected_yoy:.2f}%")
        print(f"  Diff:     {diff_yoy:.4f}")
        print(f"  Status:   {'PASS' if match_yoy else 'FAIL'}")
        
        if not match_yoy:
            mismatches.append(f"YoY % mismatch: {computed_yoy:.2f} vs {expected_yoy:.2f}")
    
    return mismatches

def generate_canonical_csv(output_path, rows):
    """Generate clean canonical CSV from valid rows."""
    print(f"\nGenerating canonical CSV: {output_path}")
    
    # Standard header
    header = ['year', 'month', 'scope', 'activity_name', 'ef_value', 'ef_unit', 
              'activity_unit', 'activity_amount', 'emission_kgco2e', 'emission_tco2e']
    
    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        
        for row in rows:
            # Map row to standard format
            # Expected: year, month, scope, activity_name, ef_value, ef_unit, activity_unit, activity_amount, emission_kgco2e, emission_tco2e
            
            # Ensure we have enough columns
            while len(row) < 10:
                row.append('')
            
            # Clean the row
            year = row[0]
            month = row[1]
            scope = row[2] if len(row) > 2 else ''
            activity = row[3] if len(row) > 3 else ''
            ef_value = row[4] if len(row) > 4 else ''
            ef_unit = row[5] if len(row) > 5 else ''
            activity_unit = row[6] if len(row) > 6 else ''
            activity_amount = row[7] if len(row) > 7 else ''
            emission_kg = row[8] if len(row) > 8 else ''
            emission_tco2e = row[9] if len(row) > 9 else ''
            
            writer.writerow([year, month, scope, activity, ef_value, ef_unit, 
                           activity_unit, activity_amount, emission_kg, emission_tco2e])
    
    print(f"Written {len(rows)} rows")

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, '../../../exdata/1.5_GreenhouseGas.xlsx')
    csv_path = os.path.join(script_dir, 'ghg_2567-2568_v1.csv')
    output_path = os.path.join(script_dir, 'ghg_2567-2568_canonical.csv')
    report_path = os.path.join(script_dir, 'ghg_validation_report.json')
    
    # Check if Excel exists
    if not os.path.exists(excel_path):
        print(f"WARNING: Excel file not found at {excel_path}")
        print("Will analyze current CSV only")
        excel_path = None
    
    # Analyze current CSV
    valid_rows, issues = analyze_current_csv(csv_path)
    
    # Compute totals from current CSV
    computed = compute_totals_from_rows(valid_rows)
    
    # Validate against ground truth
    mismatches = validate_against_ground_truth(computed)
    
    # Generate report
    report = {
        'excel_path': excel_path,
        'csv_path': csv_path,
        'valid_rows': len(valid_rows),
        'issues_count': len(issues),
        'issues_sample': issues[:20],
        'computed_totals': {str(k): float(v) for k, v in computed['year_totals'].items()},
        'mismatches': mismatches,
        'ground_truth': {
            '2567_total_tco2e': float(GROUND_TRUTH['2567_total_tco2e']),
            '2568_total_tco2e': float(GROUND_TRUTH['2568_total_tco2e']),
            'yoy_pct': float(GROUND_TRUTH['yoy_pct']),
        }
    }
    
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    
    print(f"\nReport saved to: {report_path}")
    
    # Summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Valid rows in CSV: {len(valid_rows)}")
    print(f"Issues found: {len(issues)}")
    print(f"Mismatches: {len(mismatches)}")
    
    if mismatches:
        print("\nACTION REQUIRED: Data does not match ground truth!")
        print("Please generate canonical dataset from Excel.")
    else:
        print("\nDATA VALID: All values match ground truth.")

if __name__ == '__main__':
    main()
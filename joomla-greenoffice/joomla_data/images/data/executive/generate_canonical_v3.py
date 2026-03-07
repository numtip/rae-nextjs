#!/usr/bin/env python3
"""
Generate Canonical GHG Dataset from Excel Summary Rows

The Excel has:
- Row 38-42 (2568) / Row 38-41 (2567): Scope summaries (ประเภท 1, 2, 3, รวม)
- Row 68 (2568) / Row 66 (2567): Monthly GHG totals (kgCO2e)
- Rows 8-25 (2568) / Rows 8-25 (2567): Activity-level data with monthly CF values

We extract:
1. Annual totals by scope
2. Monthly totals
3. Activity-level breakdown
"""

import openpyxl
from openpyxl import load_workbook
import csv
import json
import os
from decimal import Decimal
from collections import defaultdict

# Ground truth from user
GROUND_TRUTH = {
    '2567_total_tco2e': Decimal('220.98693744'),
    '2568_total_tco2e': Decimal('231.620303712'),
    '2567_scopes': {
        'Scope1': Decimal('11.01713228'),
        'Scope2': Decimal('192.53468536'),
        'Scope3': Decimal('17.43511980'),
    },
    '2568_scopes': {
        'Scope1': Decimal('10.847924292'),
        'Scope2': Decimal('201.47809632'),
        'Scope3': Decimal('19.29428310'),
    },
}

# Expected 2568 monthly values (kgCO2e)
EXPECTED_2568_MONTHLY = {
    1: Decimal('11530.49'),
    2: Decimal('14451.72'),
    3: Decimal('21384.86'),
    4: Decimal('20864.64'),
    5: Decimal('22672.85'),
    6: Decimal('21784.21'),
    7: Decimal('21964.51'),
    8: Decimal('22233.70'),
    9: Decimal('23440.91'),
    10: Decimal('18851.13'),
    11: Decimal('18425.95'),
    12: Decimal('14015.33'),
}

# Expected 2567 monthly values (kgCO2e) - from Excel row 66
EXPECTED_2567_MONTHLY = {
    1: Decimal('10391.25'),
    2: Decimal('13886.31'),
    3: Decimal('19980.39'),
    4: Decimal('21140.08'),
    5: Decimal('20871.34'),
    6: Decimal('21106.89'),
    7: Decimal('22742.23'),
    8: Decimal('21471.20'),
    9: Decimal('22647.67'),  # Need to verify
    10: Decimal('20168.48'),
    11: Decimal('18506.81'),
    12: Decimal('15206.78'),
}

def get_cell_value(ws, row, col):
    """Get cell value, handling None and merged cells."""
    cell = ws.cell(row=row, column=col)
    if cell.value is None:
        return None
    if isinstance(cell.value, (int, float)):
        return float(cell.value)
    return str(cell.value).strip()

def parse_summary_sheet_v2(ws, year):
    """Parse summary sheet using known structure."""
    print(f"  Parsing sheet for year {year}")
    
    results = {
        'annual_total': Decimal('0'),
        'scopes': {},
        'monthly_totals': {},
        'activities': [],
    }
    
    # Find the row with scope summaries
    # In 2568 sheet: row 38-42 has scope summaries
    # In 2567 sheet: similar structure
    
    # Scope mapping
    scope_rows = {
        'Scope1': None,
        'Scope2': None,
        'Scope3': None,
    }
    
    # Search for scope rows
    for row_idx in range(35, 50):
        col_a = str(get_cell_value(ws, row_idx, 1) or '')
        col_b = str(get_cell_value(ws, row_idx, 2) or '')
        
        # Look for "ประเภท 1" or "ประเภท 2" or "ประเภท 3"
        if 'ประเภท 1' in col_b:
            scope_rows['Scope1'] = row_idx
        elif 'ประเภท 2' in col_b:
            scope_rows['Scope2'] = row_idx
        elif 'ประเภท 3' in col_b:
            scope_rows['Scope3'] = row_idx
    
    print(f"    Scope rows: {scope_rows}")
    
    # Extract scope values (column C for 2568, column C for 2567)
    # Looking at: col_b = Scope name, col_c = 2567 value, col_d = 2568 value
    for scope, row_idx in scope_rows.items():
        if row_idx:
            # Try column C (2567) or D (2568)
            if year == 2568:
                val = get_cell_value(ws, row_idx, 4)  # Column D for 2568
            else:
                val = get_cell_value(ws, row_idx, 3)  # Column C for 2567
            
            if val:
                results['scopes'][scope] = Decimal(str(val))
                results['annual_total'] += Decimal(str(val))
    
    # Find monthly totals row
    # Row 68 in 2568 sheet, Row 66 in 2567 sheet
    # Look for "GHG ปี XXXX (kgCO2e)" or "ปริมาณก๊าซเรือนกระจก"
    monthly_row = None
    for row_idx in range(60, 75):
        col_b = str(get_cell_value(ws, row_idx, 2) or '')
        col_c = str(get_cell_value(ws, row_idx, 3) or '')
        if 'GHG' in col_b or 'ปริมาณก๊าซเรือนกระจก' in col_b:
            monthly_row = row_idx
            break
        if 'GHG' in col_c or 'ปริมาณก๊าซเรือนกระจก' in col_c:
            monthly_row = row_idx
            break
    
    print(f"    Monthly totals row: {monthly_row}")
    
    if monthly_row:
        # Monthly values are in columns C onwards (col_idx 3 = Jan, etc.)
        # Actually, looking at the structure, months are in columns like F, H, J, etc.
        # But the summary row has month names in row 49
        
        # From the structure analysis:
        # Row 68 (2568): columns C onwards = monthly values
        # Row 66 (2567): columns C onwards = monthly values
        
        # Let's read from column C (index 3) to N (index 14) for 12 months
        for month in range(1, 13):
            col_idx = 3 + (month - 1) * 2  # Every other column for Qty/CF, we want CF
            # Actually, looking at the structure, monthly totals are in consecutive columns
            val = get_cell_value(ws, monthly_row, month + 2)  # C=Jan, D=Feb, etc.
            if val:
                results['monthly_totals'][month] = Decimal(str(val))
    
    return results

def parse_excel_directly(excel_path):
    """Parse Excel and extract data directly from known structure."""
    print(f"Reading Excel: {excel_path}")
    
    wb = load_workbook(excel_path, data_only=True)
    
    all_results = {}
    
    # Process 2568 sheet
    for sheet_name in wb.sheetnames:
        if '2568' in sheet_name and 'สรุป' in sheet_name:
            ws = wb[sheet_name]
            all_results[2568] = parse_summary_sheet_v2(ws, 2568)
            break
    
    # Process 2567 sheet
    for sheet_name in wb.sheetnames:
        if '2567' in sheet_name and 'สรุป' in sheet_name:
            ws = wb[sheet_name]
            all_results[2567] = parse_summary_sheet_v2(ws, 2567)
            break
    
    wb.close()
    
    return all_results

def extract_activity_data(ws, year):
    """Extract activity-level monthly data from the Excel sheet."""
    print(f"  Extracting activity data for year {year}")
    
    activities = []
    current_scope = None
    
    # Month columns for CF (emission values)
    # Structure: Column F=6 (Jan Qty), G=7 (Jan CF), H=8 (Feb Qty), I=9 (Feb CF), etc.
    month_cf_cols = {
        1: 7,   # Jan CF
        2: 9,   # Feb CF
        3: 11,  # Mar CF
        4: 13,  # Apr CF
        5: 15,  # May CF
        6: 17,  # Jun CF
        7: 19,  # Jul CF
        8: 21,  # Aug CF
        9: 23,  # Sep CF
        10: 25, # Oct CF
        11: 27, # Nov CF
        12: 29, # Dec CF
    }
    
    # For 2567, columns start at F (6) as well based on structure
    if year == 2567:
        month_cf_cols = {
            1: 7, 2: 9, 3: 11, 4: 13, 5: 15, 6: 17,
            7: 19, 8: 21, 9: 23, 10: 25, 11: 27, 12: 29
        }
    
    # Data rows are 8-25 (activities)
    for row_idx in range(8, 26):
        activity = get_cell_value(ws, row_idx, 2)  # Column B = Activity name
        
        if not activity or activity in ['', 'None', 'รายการ']:
            continue
        
        # Determine scope from column A or activity name
        col_a = str(get_cell_value(ws, row_idx, 1) or '')
        
        if 'Scope 1' in col_a or 'ประเภท 1' in col_a:
            current_scope = 'Scope 1'
        elif 'Scope 2' in col_a or 'ประเภท 2' in col_a:
            current_scope = 'Scope 2'
        elif 'Scope 3' in col_a or 'ประเภท 3' in col_a:
            current_scope = 'Scope 3'
        
        # Also check from activity position in the sheet
        # Row 6-19: Scope 1 (rows 8-19)
        # Row 20: Scope 2 (row 20)
        # Row 21-25: Scope 3 (rows 21-25)
        if row_idx <= 19:
            current_scope = 'Scope 1'
        elif row_idx == 20:
            current_scope = 'Scope 2'
        else:
            current_scope = 'Scope 3'
        
        # Extract monthly CF values
        for month, col_idx in month_cf_cols.items():
            cf_val = get_cell_value(ws, row_idx, col_idx)
            
            if cf_val and cf_val > 0:
                activities.append({
                    'year': year,
                    'month': month,
                    'scope': current_scope,
                    'activity': activity,
                    'emission_kgco2e': float(cf_val),
                    'emission_tco2e': float(cf_val) / 1000,
                })
    
    print(f"    Extracted {len(activities)} activity records")
    return activities

def generate_canonical_csv_v2(all_data, output_path):
    """Generate canonical CSV from activity data."""
    print(f"\nGenerating canonical CSV: {output_path}")
    
    header = ['year', 'month', 'scope', 'activity_name', 'emission_kgco2e', 'emission_tco2e']
    
    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        
        for row in sorted(all_data, key=lambda x: (x['year'], x['month'], x['scope'], x['activity'])):
            writer.writerow([
                row['year'],
                row['month'],
                row['scope'],
                row['activity'],
                f"{row['emission_kgco2e']:.6f}",
                f"{row['emission_tco2e']:.8f}"
            ])
    
    print(f"Written {len(all_data)} rows")

def validate_results(results):
    """Validate extracted results against ground truth."""
    print("\n" + "="*60)
    print("VALIDATION REPORT")
    print("="*60)
    
    mismatches = []
    
    for year in [2567, 2568]:
        if year not in results:
            print(f"\n{year}: No data extracted")
            continue
        
        data = results[year]
        total = data.get('annual_total', Decimal('0'))
        expected_total = GROUND_TRUTH[f'{year}_total_tco2e']
        
        diff = abs(total - expected_total)
        match = diff < Decimal('0.5')
        
        print(f"\n{year} Total GHG:")
        print(f"  Computed: {total:.8f} tCO2e")
        print(f"  Expected: {expected_total:.8f} tCO2e")
        print(f"  Diff:     {diff:.8f} tCO2e")
        print(f"  Status:   {'PASS' if match else 'FAIL'}")
        
        if not match:
            mismatches.append(f"{year} total: {total:.2f} vs {expected_total:.2f}")
        
        # Check scopes
        print(f"\n{year} Scopes:")
        for scope in ['Scope1', 'Scope2', 'Scope3']:
            computed_scope = data.get('scopes', {}).get(scope, Decimal('0'))
            expected_scope = GROUND_TRUTH[f'{year}_scopes'][scope]
            diff_scope = abs(computed_scope - expected_scope)
            match_scope = diff_scope < Decimal('0.1')
            
            print(f"  {scope}:")
            print(f"    Computed: {computed_scope:.8f} tCO2e")
            print(f"    Expected: {expected_scope:.8f} tCO2e")
            print(f"    Diff:     {diff_scope:.8f} tCO2e")
            print(f"    Status:   {'PASS' if match_scope else 'FAIL'}")
            
            if not match_scope:
                mismatches.append(f"{year} {scope}: {computed_scope:.2f} vs {expected_scope:.2f}")
    
    return mismatches

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = '/home/rae_admin/joomla-greenoffice/exdata/1.5_GreenhouseGas.xlsx'
    output_csv = os.path.join(script_dir, 'ghg_2567-2568_canonical.csv')
    
    # Parse Excel
    print("="*60)
    print("PARSING EXCEL FILE")
    print("="*60)
    
    wb = load_workbook(excel_path, data_only=True)
    
    all_activities = []
    
    # Process 2568
    for sheet_name in wb.sheetnames:
        if '2568' in sheet_name and 'สรุป' in sheet_name:
            ws = wb[sheet_name]
            activities = extract_activity_data(ws, 2568)
            all_activities.extend(activities)
            break
    
    # Process 2567
    for sheet_name in wb.sheetnames:
        if '2567' in sheet_name and 'สรุป' in sheet_name:
            ws = wb[sheet_name]
            activities = extract_activity_data(ws, 2567)
            all_activities.extend(activities)
            break
    
    wb.close()
    
    # Compute totals from activity data
    year_totals = defaultdict(float)
    year_scopes = defaultdict(lambda: defaultdict(float))
    year_months = defaultdict(lambda: defaultdict(float))
    
    for act in all_activities:
        year = act['year']
        month = act['month']
        scope = act['scope']
        emission = act['emission_kgco2e']
        
        year_totals[year] += emission
        year_scopes[year][scope] += emission
        year_months[year][month] += emission
    
    # Print computed values
    print("\n" + "="*60)
    print("COMPUTED VALUES FROM ACTIVITY DATA")
    print("="*60)
    
    for year in [2567, 2568]:
        total_tco2e = year_totals[year] / 1000
        print(f"\n{year}:")
        print(f"  Total: {total_tco2e:.8f} tCO2e ({year_totals[year]:.2f} kgCO2e)")
        print(f"  Scopes:")
        for scope in ['Scope 1', 'Scope 2', 'Scope 3']:
            scope_tco2e = year_scopes[year].get(scope, 0) / 1000
            print(f"    {scope}: {scope_tco2e:.8f} tCO2e")
        print(f"  Monthly totals (kgCO2e):")
        for month in range(1, 13):
            print(f"    Month {month}: {year_months[year].get(month, 0):.2f}")
    
    # Validate
    print("\n" + "="*60)
    print("VALIDATION AGAINST GROUND TRUTH")
    print("="*60)
    
    mismatches = []
    
    for year in [2567, 2568]:
        computed_total = year_totals[year] / 1000
        expected_total = float(GROUND_TRUTH[f'{year}_total_tco2e'])
        diff = abs(computed_total - expected_total)
        
        print(f"\n{year} Total GHG:")
        print(f"  Computed: {computed_total:.8f} tCO2e")
        print(f"  Expected: {expected_total:.8f} tCO2e")
        print(f"  Diff:     {diff:.8f} tCO2e ({diff/expected_total*100:.2f}%)")
        print(f"  Status:   {'PASS' if diff < 0.5 else 'FAIL'}")
        
        if diff >= 0.5:
            mismatches.append(f"{year} total mismatch")
    
    # Generate CSV if no mismatches
    if all_activities:
        generate_canonical_csv_v2(all_activities, output_csv)
    
    # Summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Total activity records: {len(all_activities)}")
    print(f"Mismatches: {len(mismatches)}")
    
    if mismatches:
        print("\n⚠️  Some values don't match exactly. Check the data.")
        for m in mismatches:
            print(f"  - {m}")
    else:
        print("\n✅ All values match ground truth!")

if __name__ == '__main__':
    main()
#!/usr/bin/env python3
"""
Generate Canonical GHG Dataset from Excel - Fixed Version

CORRECT Column Mapping:
- Row 4: Month names (ม.ค., ก.พ., ...)
- Row 5: Sub-headers (ปริมาณ, CF)
- Row 8-25: Activity data

For each month, there are 2 columns:
- Column N = Qty (ปริมาณ)
- Column N+1 = CF (emission in kgCO2e)

Correct CF columns for 12 months: 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30
"""

import openpyxl
from openpyxl import load_workbook
import csv
import json
import os
from decimal import Decimal
from collections import defaultdict

# Ground truth from user
GROUND_TRUTH = {
    '2567_total_tco2e': Decimal('220.98693744'),
    '2568_total_tco2e': Decimal('231.620303712'),
    '2567_scopes': {
        'Scope1': Decimal('11.01713228'),
        'Scope2': Decimal('192.53468536'),
        'Scope3': Decimal('17.43511980'),
    },
    '2568_scopes': {
        'Scope1': Decimal('10.847924292'),
        'Scope2': Decimal('201.47809632'),
        'Scope3': Decimal('19.29428310'),
    },
}

# Expected 2568 monthly values (kgCO2e)
EXPECTED_2568_MONTHLY = {
    1: Decimal('11530.49'),
    2: Decimal('14451.72'),
    3: Decimal('21384.86'),
    4: Decimal('20864.64'),
    5: Decimal('22672.85'),
    6: Decimal('21784.21'),
    7: Decimal('21964.51'),
    8: Decimal('22233.70'),
    9: Decimal('23440.91'),
    10: Decimal('18851.13'),
    11: Decimal('18425.95'),
    12: Decimal('14015.33'),
}

def get_cell_value(ws, row, col):
    """Get cell value, handling None and merged cells."""
    cell = ws.cell(row=row, column=col)
    if cell.value is None:
        return None
    if isinstance(cell.value, (int, float)):
        return float(cell.value)
    return str(cell.value).strip()

def extract_activity_data_fixed(ws, year):
    """Extract activity data with CORRECT column mapping."""
    print(f"  Extracting activity data for year {year}")
    
    activities = []
    current_scope = None
    
    # CORRECT: CF columns are at positions 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30
    # Because: Col 7 = Jan Qty, Col 8 = Jan CF, Col 9 = Feb Qty, Col 10 = Feb CF, etc.
    month_cf_cols = {
        1: 8,   # Jan CF
        2: 10,  # Feb CF
        3: 12,  # Mar CF
        4: 14,  # Apr CF
        5: 16,  # May CF
        6: 18,  # Jun CF
        7: 20,  # Jul CF
        8: 22,  # Aug CF
        9: 24,  # Sep CF
        10: 26, # Oct CF
        11: 28, # Nov CF
        12: 30, # Dec CF
    }
    
    # Data rows: 8-25 (activities)
    # Row 8-19: Scope 1 activities
    # Row 20: Scope 2 (electricity)
    # Row 21-25: Scope 3 activities
    
    for row_idx in range(8, 26):
        activity = get_cell_value(ws, row_idx, 2)  # Column B = Activity name
        
        if not activity or activity in ['', 'None', 'รายการ']:
            continue
        
        # Determine scope from row position
        if row_idx <= 19:
            current_scope = 'Scope 1'
        elif row_idx == 20:
            current_scope = 'Scope 2'
        else:
            current_scope = 'Scope 3'
        
        # Also check column A for scope marker
        col_a = str(get_cell_value(ws, row_idx, 1) or '')
        if 'Scope 1' in col_a or 'ประเภท 1' in col_a:
            current_scope = 'Scope 1'
        elif 'Scope 2' in col_a or 'ประเภท 2' in col_a:
            current_scope = 'Scope 2'
        elif 'Scope 3' in col_a or 'ประเภท 3' in col_a:
            current_scope = 'Scope 3'
        
        # Extract monthly CF values
        for month, col_idx in month_cf_cols.items():
            cf_val = get_cell_value(ws, row_idx, col_idx)
            
            if cf_val is not None and cf_val != '' and cf_val > 0:
                activities.append({
                    'year': year,
                    'month': month,
                    'scope': current_scope,
                    'activity': activity,
                    'emission_kgco2e': float(cf_val),
                    'emission_tco2e': float(cf_val) / 1000,
                })
    
    print(f"    Extracted {len(activities)} activity records")
    return activities

def generate_canonical_csv(all_data, output_path):
    """Generate canonical CSV."""
    print(f"\nGenerating canonical CSV: {output_path}")
    
    header = ['year', 'month', 'scope', 'activity_name', 'emission_kgco2e', 'emission_tco2e']
    
    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        
        for row in sorted(all_data, key=lambda x: (x['year'], x['month'], x['scope'], x['activity'])):
            writer.writerow([
                row['year'],
                row['month'],
                row['scope'],
                row['activity'],
                f"{row['emission_kgco2e']:.6f}",
                f"{row['emission_tco2e']:.8f}"
            ])
    
    print(f"Written {len(all_data)} rows")

def validate_results(all_activities):
    """Validate extracted results against ground truth."""
    print("\n" + "="*60)
    print("VALIDATION REPORT")
    print("="*60)
    
    # Compute totals from activities
    year_totals = defaultdict(float)
    year_scopes = defaultdict(lambda: defaultdict(float))
    year_months = defaultdict(lambda: defaultdict(float))
    
    for act in all_activities:
        year = act['year']
        month = act['month']
        scope = act['scope']
        emission = act['emission_kgco2e']
        
        year_totals[year] += emission
        year_scopes[year][scope] += emission
        year_months[year][month] += emission
    
    mismatches = []
    
    for year in [2567, 2568]:
        total_kg = year_totals[year]
        total_tco2e = total_kg / 1000
        
        expected_total = float(GROUND_TRUTH[f'{year}_total_tco2e'])
        diff = abs(total_tco2e - expected_total)
        match = diff < 0.5
        
        print(f"\n{year} Total GHG:")
        print(f"  Computed: {total_tco2e:.8f} tCO2e ({total_kg:.2f} kgCO2e)")
        print(f"  Expected: {expected_total:.8f} tCO2e")
        print(f"  Diff:     {diff:.8f} tCO2e ({diff/expected_total*100:.2f}%)")
        print(f"  Status:   {'✅ PASS' if match else '❌ FAIL'}")
        
        if not match:
            mismatches.append(f"{year} total mismatch: {total_tco2e:.2f} vs {expected_total:.2f}")
        
        # Check scopes
        print(f"\n{year} Scopes:")
        for scope in ['Scope 1', 'Scope 2', 'Scope 3']:
            scope_kg = year_scopes[year].get(scope, 0)
            scope_tco2e = scope_kg / 1000
            
            expected_scope = float(GROUND_TRUTH[f'{year}_scopes'][scope.replace(' ', '')])
            diff_scope = abs(scope_tco2e - expected_scope)
            match_scope = diff_scope < 0.1
            
            print(f"  {scope}:")
            print(f"    Computed: {scope_tco2e:.8f} tCO2e")
            print(f"    Expected: {expected_scope:.8f} tCO2e")
            print(f"    Diff:     {diff_scope:.8f} tCO2e")
            print(f"    Status:   {'✅ PASS' if match_scope else '❌ FAIL'}")
            
            if not match_scope:
                mismatches.append(f"{year} {scope} mismatch")
        
        # Check monthly totals for 2568
        if year == 2568:
            print(f"\n{year} Monthly Totals (kgCO2e):")
            for month in range(1, 13):
                computed_kg = year_months[year].get(month, 0)
                expected_kg = float(EXPECTED_2568_MONTHLY.get(month, 0))
                diff_month = abs(computed_kg - expected_kg)
                match_month = diff_month < 1.0
                
                print(f"    Month {month:2}: Computed={computed_kg:.2f}, Expected={expected_kg:.2f}, Diff={diff_month:.2f} {'✅' if match_month else '❌'}")
                
                if not match_month:
                    mismatches.append(f"{year} month {month} mismatch")
    
    return mismatches, year_totals, year_scopes, year_months

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = '/home/rae_admin/joomla-greenoffice/exdata/1.5_GreenhouseGas.xlsx'
    output_csv = os.path.join(script_dir, 'ghg_2567-2568_canonical.csv')
    
    print("="*60)
    print("PARSING EXCEL FILE (FIXED COLUMN MAPPING)")
    print("="*60)
    
    wb = load_workbook(excel_path, data_only=True)
    
    all_activities = []
    
    # Process 2568
    for sheet_name in wb.sheetnames:
        if '2568' in sheet_name and 'สรุป' in sheet_name:
            ws = wb[sheet_name]
            activities = extract_activity_data_fixed(ws, 2568)
            all_activities.extend(activities)
            break
    
    # Process 2567
    for sheet_name in wb.sheetnames:
        if '2567' in sheet_name and 'สรุป' in sheet_name:
            ws = wb[sheet_name]
            activities = extract_activity_data_fixed(ws, 2567)
            all_activities.extend(activities)
            break
    
    wb.close()
    
    # Validate
    mismatches, year_totals, year_scopes, year_months = validate_results(all_activities)
    
    # Generate CSV
    if all_activities:
        generate_canonical_csv(all_activities, output_csv)
    
    # Summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Total activity records: {len(all_activities)}")
    print(f"Mismatches: {len(mismatches)}")
    
    if mismatches:
        print("\n⚠️  Some values don't match:")
        for m in mismatches:
            print(f"  - {m}")
    else:
        print("\n✅ ALL VALUES MATCH GROUND TRUTH!")
        print("\nCanonical CSV generated successfully.")
        print(f"Path: {output_csv}")
        
        # Print summary statistics
        print("\nFinal Statistics:")
        for year in [2567, 2568]:
            total = year_totals[year] / 1000
            print(f"  {year}: {total:.2f} tCO2e")

if __name__ == '__main__':
    main()
#!/usr/bin/env python3
"""
Generate Canonical GHG Dataset from Excel Source

This script:
1. Reads the Excel file as ground truth
2. Extracts all GHG emission data for 2567 and 2568
3. Generates a clean CSV for the dashboard
4. Validates against expected totals

Ground truth values from Excel:
- 2567 total GHG = 220.98693744 tCO2e
- 2568 total GHG = 231.620303712 tCO2e
- YoY total = +10.633366272 tCO2e = +4.81%
"""

import openpyxl
from openpyxl import load_workbook
import csv
import json
import os
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict

# Ground truth from Excel (user provided)
GROUND_TRUTH = {
    '2567_total_tco2e': Decimal('220.98693744'),
    '2568_total_tco2e': Decimal('231.620303712'),
    'yoy_diff_tco2e': Decimal('10.633366272'),
    'yoy_pct': Decimal('4.81'),
}

# Expected monthly values for 2568 (in kgCO2e)
EXPECTED_2568_MONTHLY_KGCO2E = {
    1: Decimal('11530.49'),
    2: Decimal('14451.72'),
    3: Decimal('21384.86'),
    4: Decimal('20864.64'),
    5: Decimal('22672.85'),
    6: Decimal('21784.21'),
    7: Decimal('21964.51'),
    8: Decimal('22233.70'),
    9: Decimal('23440.91'),
    10: Decimal('18851.13'),
    11: Decimal('18425.95'),
    12: Decimal('14015.33'),
}

# Expected 2568 scopes
EXPECTED_2568_SCOPES = {
    'Scope1': Decimal('10.847924292'),
    'Scope2': Decimal('201.47809632'),
    'Scope3': Decimal('19.29428310'),
}

# Expected 2567 scopes
EXPECTED_2567_SCOPES = {
    'Scope1': Decimal('11.01713228'),
    'Scope2': Decimal('192.53468536'),
    'Scope3': Decimal('17.43511980'),
}

# Expected 2568 top activities
EXPECTED_2568_TOP_ACTIVITIES = [
    ('electricity', Decimal('201.47809632')),
    ('landfill waste', Decimal('10.16392')),
    ('septic tank CH4', Decimal('7.8204')),
    ('paper', Decimal('4.6197756')),
    ('provincial water', Decimal('4.5105875')),
]

def normalize_scope(scope_str):
    """Normalize scope string to standard format."""
    if not scope_str:
        return ''
    s = str(scope_str).strip().lower()
    if 'scope 1' in s or 'scope1' in s or 'ประเภท 1' in s:
        return 'Scope 1'
    elif 'scope 2' in s or 'scope2' in s or 'ประเภท 2' in s:
        return 'Scope 2'
    elif 'scope 3' in s or 'scope3' in s or 'ประเภท 3' in s:
        return 'Scope 3'
    return scope_str.strip()

def extract_excel_data(excel_path):
    """Extract clean GHG data from Excel file."""
    print(f"Reading Excel: {excel_path}")
    
    wb = load_workbook(excel_path, data_only=True)
    
    all_data = []
    year_data = defaultdict(lambda: {
        'total_kgco2e': Decimal('0'),
        'scopes': defaultdict(lambda: Decimal('0')),
        'activities': defaultdict(lambda: Decimal('0')),
        'months': defaultdict(lambda: Decimal('0')),
    })
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Processing sheet: {sheet_name}")
        
        # Find header row
        header = None
        header_row_idx = None
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_values = [str(c).strip() if c else '' for c in row]
            # Look for year column
            if any(v and ('year' in str(v).lower() or 'ปี' in str(v)) for v in row_values):
                header = row_values
                header_row_idx = row_idx
                break
        
        if not header:
            print(f"    No header found in sheet {sheet_name}")
            continue
        
        print(f"    Header at row {header_row_idx}: {header[:8]}...")
        
        # Process data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), header_row_idx + 1):
            row_values = []
            for cell in row:
                if cell is None:
                    row_values.append('')
                elif isinstance(cell, (int, float)):
                    row_values.append(cell)
                else:
                    row_values.append(str(cell).strip())
            
            # Skip non-data rows
            if not row_values or len(row_values) < 3:
                continue
            
            # Check for year
            first_col = row_values[0]
            if isinstance(first_col, (int, float)):
                year = int(first_col)
            else:
                try:
                    year = int(float(str(first_col)))
                except (ValueError, TypeError):
                    continue
            
            if year not in (2567, 2568):
                continue
            
            # Extract data based on column structure
            # We need: year, month, scope, activity, emission_kgco2e
            # The Excel might have different structures
            
            # Try to find month column
            month = None
            for i, h in enumerate(header[:5]):
                if h and ('month' in str(h).lower() or 'เดือน' in str(h)):
                    if i < len(row_values):
                        month = row_values[i]
                        if isinstance(month, (int, float)):
                            month = int(month)
                        else:
                            try:
                                month = int(float(str(month)))
                            except:
                                month = None
                    break
            
            # Try to find scope column
            scope = None
            for i, h in enumerate(header[:8]):
                if h and ('scope' in str(h).lower() or 'ประเภท' in str(h)):
                    if i < len(row_values):
                        scope = normalize_scope(row_values[i])
                    break
            
            # If scope not found in header, look for it in data
            if scope is None and len(row_values) > 2:
                scope = normalize_scope(row_values[2])
            
            # Try to find activity column
            activity = None
            for i, h in enumerate(header[:8]):
                if h and ('activity' in str(h).lower() or 'รายการ' in str(h) or 'กิจกรรม' in str(h)):
                    if i < len(row_values):
                        activity = row_values[i]
                    break
            
            # If activity not found, try column index
            if activity is None and len(row_values) > 3:
                activity = str(row_values[3]) if row_values[3] else ''
            
            # Try to find emission column (kgCO2e or tCO2e)
            emission_kg = None
            emission_t = None
            
            for i, h in enumerate(header):
                h_str = str(h).lower() if h else ''
                if 'kgco2e' in h_str or 'kg co2e' in h_str:
                    if i < len(row_values):
                        val = row_values[i]
                        if isinstance(val, (int, float)):
                            emission_kg = Decimal(str(val))
                        elif val:
                            try:
                                emission_kg = Decimal(str(val).replace(',', ''))
                            except:
                                pass
                elif 'tco2e' in h_str or 't co2e' in h_str:
                    if i < len(row_values):
                        val = row_values[i]
                        if isinstance(val, (int, float)):
                            emission_t = Decimal(str(val))
                        elif val:
                            try:
                                emission_t = Decimal(str(val).replace(',', ''))
                            except:
                                pass
            
            # If emission values found, accumulate
            if emission_kg is not None or emission_t is not None:
                # Use kgCO2e if available, else convert from tCO2e
                if emission_kg is not None:
                    emission = emission_kg
                else:
                    emission = emission_t * 1000 if emission_t else Decimal('0')
                
                if year and month and scope:
                    year_data[year]['total_kgco2e'] += emission
                    year_data[year]['scopes'][scope] += emission
                    if activity:
                        year_data[year]['activities'][activity] += emission
                    if month:
                        year_data[year]['months'][month] += emission
                    
                    all_data.append({
                        'year': year,
                        'month': month,
                        'scope': scope,
                        'activity': activity or '',
                        'emission_kgco2e': float(emission),
                        'emission_tco2e': float(emission / 1000),
                    })
    
    wb.close()
    
    print(f"\nExtracted {len(all_data)} data rows")
    
    # Print summary
    for year in sorted(year_data.keys()):
        data = year_data[year]
        print(f"\nYear {year}:")
        print(f"  Total: {data['total_kgco2e']/1000:.8f} tCO2e")
        print(f"  Scopes: {dict((k, float(v/1000)) for k, v in data['scopes'].items())}")
        print(f"  Months: {len(data['months'])} months")
    
    return all_data, year_data

def generate_canonical_csv(data, output_path):
    """Generate clean canonical CSV."""
    print(f"\nGenerating canonical CSV: {output_path}")
    
    header = ['year', 'month', 'scope', 'activity_name', 'emission_kgco2e', 'emission_tco2e']
    
    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        
        for row in sorted(data, key=lambda x: (x['year'], x['month'], x['scope'])):
            writer.writerow([
                row['year'],
                row['month'],
                row['scope'],
                row['activity'],
                row['emission_kgco2e'],
                row['emission_tco2e']
            ])
    
    print(f"Written {len(data)} rows")

def validate_against_ground_truth(year_data):
    """Validate computed values against ground truth."""
    print("\n" + "="*60)
    print("VALIDATION AGAINST GROUND TRUTH")
    print("="*60)
    
    mismatches = []
    
    # 2567 total
    computed_2567 = year_data[2567]['total_kgco2e'] / 1000
    expected_2567 = GROUND_TRUTH['2567_total_tco2e']
    diff_2567 = abs(computed_2567 - expected_2567)
    match_2567 = diff_2567 < Decimal('0.5')
    
    print(f"\n2567 Total GHG:")
    print(f"  Computed: {computed_2567:.8f} tCO2e")
    print(f"  Expected: {expected_2567:.8f} tCO2e")
    print(f"  Diff:     {diff_2567:.8f} tCO2e ({float(diff_2567/expected_2567*100):.2f}%)")
    print(f"  Status:   {'PASS' if match_2567 else 'FAIL'}")
    
    if not match_2567:
        mismatches.append(f"2567 total: {computed_2567:.2f} vs {expected_2567:.2f}")
    
    # 2568 total
    computed_2568 = year_data[2568]['total_kgco2e'] / 1000
    expected_2568 = GROUND_TRUTH['2568_total_tco2e']
    diff_2568 = abs(computed_2568 - expected_2568)
    match_2568 = diff_2568 < Decimal('0.5')
    
    print(f"\n2568 Total GHG:")
    print(f"  Computed: {computed_2568:.8f} tCO2e")
    print(f"  Expected: {expected_2568:.8f} tCO2e")
    print(f"  Diff:     {diff_2568:.8f} tCO2e ({float(diff_2568/expected_2568*100):.2f}%)")
    print(f"  Status:   {'PASS' if match_2568 else 'FAIL'}")
    
    if not match_2568:
        mismatches.append(f"2568 total: {computed_2568:.2f} vs {expected_2568:.2f}")
    
    # YoY
    if computed_2567 > 0:
        computed_yoy = float((computed_2568 - computed_2567) / computed_2567 * 100)
        expected_yoy = float(GROUND_TRUTH['yoy_pct'])
        diff_yoy = abs(computed_yoy - expected_yoy)
        match_yoy = diff_yoy < 0.5
        
        print(f"\nYoY %:")
        print(f"  Computed: {computed_yoy:.2f}%")
        print(f"  Expected: {expected_yoy:.2f}%")
        print(f"  Diff:     {diff_yoy:.4f}")
        print(f"  Status:   {'PASS' if match_yoy else 'FAIL'}")
        
        if not match_yoy:
            mismatches.append(f"YoY %: {computed_yoy:.2f} vs {expected_yoy:.2f}")
    
    # Check 2568 scopes
    print(f"\n2568 Scopes:")
    for scope, expected in EXPECTED_2568_SCOPES.items():
        computed = year_data[2568]['scopes'].get(scope, Decimal('0')) / 1000
        diff = abs(computed - expected)
        match = diff < Decimal('0.1')
        print(f"  {scope}:")
        print(f"    Computed: {computed:.8f} tCO2e")
        print(f"    Expected: {expected:.8f} tCO2e")
        print(f"    Diff:     {diff:.8f} tCO2e")
        print(f"    Status:   {'PASS' if match else 'FAIL'}")
        if not match:
            mismatches.append(f"2568 {scope}: {computed:.2f} vs {expected:.2f}")
    
    # Check 2567 scopes
    print(f"\n2567 Scopes:")
    for scope, expected in EXPECTED_2567_SCOPES.items():
        computed = year_data[2567]['scopes'].get(scope, Decimal('0')) / 1000
        diff = abs(computed - expected)
        match = diff < Decimal('0.1')
        print(f"  {scope}:")
        print(f"    Computed: {computed:.8f} tCO2e")
        print(f"    Expected: {expected:.8f} tCO2e")
        print(f"    Diff:     {diff:.8f} tCO2e")
        print(f"    Status:   {'PASS' if match else 'FAIL'}")
        if not match:
            mismatches.append(f"2567 {scope}: {computed:.2f} vs {expected:.2f}")
    
    return mismatches

def analyze_excel_structure(excel_path):
    """Analyze the structure of the Excel file."""
    print(f"\nAnalyzing Excel structure: {excel_path}")
    
    wb = load_workbook(excel_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== Sheet: {sheet_name} ===")
        print(f"Dimensions: {ws.dimensions}")
        
        # Print first 15 rows to understand structure
        print("\nFirst 15 rows:")
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
            row_str = []
            for i, cell in enumerate(row[:12]):  # First 12 columns
                if cell is None:
                    row_str.append('')
                elif isinstance(cell, float):
                    row_str.append(f"{cell:.2f}")
                else:
                    row_str.append(str(cell)[:30])
            print(f"Row {row_idx:2}: {' | '.join(row_str)}")
    
    wb.close()

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = '/home/rae_admin/joomla-greenoffice/exdata/1.5_GreenhouseGas.xlsx'
    output_csv = os.path.join(script_dir, 'ghg_2567-2568_canonical.csv')
    output_json = os.path.join(script_dir, 'ghg_canonical_summary.json')
    
    # First, analyze structure
    print("="*60)
    print("ANALYZING EXCEL STRUCTURE")
    print("="*60)
    analyze_excel_structure(excel_path)
    
    # Extract data
    print("\n" + "="*60)
    print("EXTRACTING DATA FROM EXCEL")
    print("="*60)
    all_data, year_data = extract_excel_data(excel_path)
    
    # Validate
    mismatches = validate_against_ground_truth(year_data)
    
    # Generate canonical CSV
    if all_data:
        generate_canonical_csv(all_data, output_csv)
    
    # Save summary
    summary = {
        'total_rows': len(all_data),
        'years': list(year_data.keys()),
        'year_totals': {str(y): float(d['total_kgco2e']/1000) for y, d in year_data.items()},
        'year_scopes': {str(y): {k: float(v/1000) for k, v in d['scopes'].items()} for y, d in year_data.items()},
        'mismatches': mismatches,
    }
    
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)
    
    print(f"\nSummary saved to: {output_json}")
    
    # Final status
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Total rows: {len(all_data)}")
    print(f"Mismatches: {len(mismatches)}")
    
    if mismatches:
        print("\n⚠️  DATA MISMATCH DETECTED!")
        print("Please review the extraction logic.")
    else:
        print("\n✅ ALL VALUES MATCH GROUND TRUTH")

if __name__ == '__main__':
    main()
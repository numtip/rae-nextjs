#!/usr/bin/env python3
"""
Parse GHG Excel file with grid layout understanding.

The Excel has:
- Rows 1-2: Headers
- Row 3: Scope/Activity headers  
- Row 4: Month names (ม.ค., ก.พ., ... for Jan-Dec) or year headers
- Row 5: Sub-headers (ปริมาณ, CF for each month)
- Row 6+: Data rows with activities

Each activity row has:
- Column A: Scope name (repeated or merged)
- Column B: Activity name
- Columns F onwards: Monthly values (alternating ปริมาณ/CF)

We need to extract the CF (emission in kgCO2e) for each month.
"""

import openpyxl
from openpyxl import load_workbook
import csv
import json
import os
from decimal import Decimal
from collections import defaultdict

# Month mapping (Thai to number)
THAI_MONTHS = {
    'ม.ค.': 1, 'มกราคม': 1,
    'ก.พ.': 2, 'กุมภาพันธ์': 2,
    'มี.ค.': 3, 'มีนาคม': 3,
    'เม.ย.': 4, 'เมษายน': 4,
    'พ.ค.': 5, 'พฤษภาคม': 5,
    'มิ.ย.': 6, 'มิถุนายน': 6,
    'ก.ค.': 7, 'กรกฎาคม': 7,
    'ส.ค.': 8, 'สิงหาคม': 8,
    'ก.ย.': 9, 'กันยายน': 9,
    'ต.ค.': 10, 'ตุลาคม': 10,
    'พ.ย.': 11, 'พฤศจิกายน': 11,
    'ธ.ค.': 12, 'ธันวาคม': 12,
}

# Ground truth from user
GROUND_TRUTH = {
    '2567_total_tco2e': Decimal('220.98693744'),
    '2568_total_tco2e': Decimal('231.620303712'),
    '2567_scopes': {
        'Scope1': Decimal('11.01713228'),
        'Scope2': Decimal('192.53468536'),
        'Scope3': Decimal('17.43511980'),
    },
    '2568_scopes': {
        'Scope1': Decimal('10.847924292'),
        'Scope2': Decimal('201.47809632'),
        'Scope3': Decimal('19.29428310'),
    },
}

def get_cell_value(ws, row, col):
    """Get cell value, handling None and merged cells."""
    cell = ws.cell(row=row, column=col)
    if cell.value is None:
        return None
    if isinstance(cell.value, (int, float)):
        return cell.value
    return str(cell.value).strip()

def parse_summary_sheet(ws, year):
    """Parse a summary sheet and extract GHG data."""
    print(f"  Parsing sheet for year {year}")
    
    data_rows = []
    current_scope = None
    
    # Find the structure
    # Row 3 typically has "ขอบเขตการดำเนินงาน" in column A
    # Row 4 has month names
    
    # Find column indices for months
    # The structure is: [Scope, Activity, EF, Unit, Unit2, Jan_Qty, Jan_CF, Feb_Qty, Feb_CF, ...]
    
    month_cols = {}  # month -> (qty_col, cf_col)
    header_row = None
    
    # Look for month headers in rows 3-5
    for row_idx in range(1, 6):
        for col_idx in range(1, 30):  # Check first 30 columns
            val = get_cell_value(ws, row_idx, col_idx)
            if val:
                val_lower = val.lower()
                for thai_month, month_num in THAI_MONTHS.items():
                    if thai_month in val:
                        # Found a month column
                        # Next column might be CF
                        month_cols[month_num] = (col_idx, col_idx + 1)
                        header_row = row_idx
                        break
    
    print(f"    Found month columns: {month_cols}")
    print(f"    Header row: {header_row}")
    
    # If no Thai months found, try English month names
    if not month_cols:
        eng_months = {
            'jan': 1, 'january': 1,
            'feb': 2, 'february': 2,
            'mar': 3, 'march': 3,
            'apr': 4, 'april': 4,
            'may': 5,
            'jun': 6, 'june': 6,
            'jul': 7, 'july': 7,
            'aug': 8, 'august': 8,
            'sep': 9, 'september': 9,
            'oct': 10, 'october': 10,
            'nov': 11, 'november': 11,
            'dec': 12, 'december': 12,
        }
        for row_idx in range(1, 6):
            for col_idx in range(1, 30):
                val = get_cell_value(ws, row_idx, col_idx)
                if val:
                    val_lower = val.lower()
                    for eng_month, month_num in eng_months.items():
                        if eng_month in val_lower:
                            month_cols[month_num] = (col_idx, col_idx + 1)
                            header_row = row_idx
                            break
    
    # If still no months, assume standard column positions
    if not month_cols:
        print("    No month columns found, using standard positions (F onwards)")
        # Standard position: F=6 is Jan Qty, G=7 is Jan CF, etc.
        for month in range(1, 13):
            qty_col = 6 + (month - 1) * 2
            cf_col = qty_col + 1
            month_cols[month] = (qty_col, cf_col)
    
    # Now extract data rows
    # Data typically starts around row 6-8
    
    for row_idx in range(6, ws.max_row + 1):
        scope = get_cell_value(ws, row_idx, 1)  # Column A
        activity = get_cell_value(ws, row_idx, 2)  # Column B
        
        # Skip empty rows or summary rows
        if not activity or activity in ['', 'None', 'รายการ']:
            continue
        
        # Determine scope from the first column or tracking
        if scope:
            if 'Scope 1' in str(scope) or 'ประเภท 1' in str(scope):
                current_scope = 'Scope 1'
            elif 'Scope 2' in str(scope) or 'ประเภท 2' in str(scope):
                current_scope = 'Scope 2'
            elif 'Scope 3' in str(scope) or 'ประเภท 3' in str(scope):
                current_scope = 'Scope 3'
        
        # Get monthly emission values
        for month in range(1, 13):
            if month not in month_cols:
                continue
            
            cf_col = month_cols[month][1]  # Use CF column (emission value)
            cf_val = get_cell_value(ws, row_idx, cf_col)
            
            if cf_val is not None and cf_val != '' and cf_val != 0:
                try:
                    emission = float(cf_val)
                    if emission > 0:
                        data_rows.append({
                            'year': year,
                            'month': month,
                            'scope': current_scope or 'Unknown',
                            'activity': activity,
                            'emission_kgco2e': emission,
                            'emission_tco2e': emission / 1000,
                        })
                except (ValueError, TypeError):
                    pass
    
    print(f"    Extracted {len(data_rows)} data rows")
    
    return data_rows

def parse_excel_v2(excel_path):
    """Parse Excel file with improved understanding of structure."""
    print(f"Reading Excel: {excel_path}")
    
    wb = load_workbook(excel_path, data_only=True)
    
    all_data = []
    
    # Process 2568 sheet
    for sheet_name in wb.sheetnames:
        if '2568' in sheet_name and 'สรุป' in sheet_name:
            ws = wb[sheet_name]
            data = parse_summary_sheet(ws, 2568)
            all_data.extend(data)
            break
    
    # Process 2567 sheet
    for sheet_name in wb.sheetnames:
        if '2567' in sheet_name and 'สรุป' in sheet_name:
            ws = wb[sheet_name]
            data = parse_summary_sheet(ws, 2567)
            all_data.extend(data)
            break
    
    wb.close()
    
    return all_data

def compute_totals(all_data):
    """Compute totals from extracted data."""
    year_totals = defaultdict(float)
    year_scopes = defaultdict(lambda: defaultdict(float))
    year_months = defaultdict(lambda: defaultdict(float))
    year_activities = defaultdict(lambda: defaultdict(float))
    
    for row in all_data:
        year = row['year']
        month = row['month']
        scope = row['scope']
        activity = row['activity']
        emission = row['emission_kgco2e']
        
        year_totals[year] += emission
        year_scopes[year][scope] += emission
        year_months[year][month] += emission
        year_activities[year][activity] += emission
    
    return {
        'year_totals': dict(year_totals),
        'year_scopes': {y: dict(s) for y, s in year_scopes.items()},
        'year_months': {y: dict(m) for y, m in year_months.items()},
        'year_activities': {y: dict(a) for y, a in year_activities.items()},
    }

def validate_and_report(computed):
    """Validate computed values and print report."""
    print("\n" + "="*60)
    print("VALIDATION REPORT")
    print("="*60)
    
    mismatches = []
    
    for year in [2567, 2568]:
        total_kg = computed['year_totals'].get(year, 0)
        total_tco2e = total_kg / 1000
        
        expected_key = f'{year}_total_tco2e'
        expected = GROUND_TRUTH[expected_key]
        
        diff = abs(total_tco2e - float(expected))
        match = diff < 0.5
        
        print(f"\n{year} Total GHG:")
        print(f"  Computed: {total_tco2e:.8f} tCO2e")
        print(f"  Expected: {expected:.8f} tCO2e")
        print(f"  Diff:     {diff:.8f} tCO2e")
        print(f"  Status:   {'PASS' if match else 'FAIL'}")
        
        if not match:
            mismatches.append(f"{year} total: {total_tco2e:.2f} vs {expected:.2f}")
        
        # Check scopes
        scope_key = f'{year}_scopes'
        print(f"\n{year} Scopes:")
        for scope in ['Scope1', 'Scope2', 'Scope3']:
            computed_scope = computed['year_scopes'].get(year, {}).get(scope, 0) / 1000
            expected_scope = GROUND_TRUTH[scope_key][scope]
            diff_scope = abs(computed_scope - float(expected_scope))
            match_scope = diff_scope < 0.1
            
            print(f"  {scope}:")
            print(f"    Computed: {computed_scope:.8f} tCO2e")
            print(f"    Expected: {expected_scope:.8f} tCO2e")
            print(f"    Diff:     {diff_scope:.8f} tCO2e")
            print(f"    Status:   {'PASS' if match_scope else 'FAIL'}")
            
            if not match_scope:
                mismatches.append(f"{year} {scope}: {computed_scope:.2f} vs {expected_scope:.2f}")
    
    return mismatches

def generate_canonical_csv(all_data, output_path):
    """Generate canonical CSV from extracted data."""
    print(f"\nGenerating canonical CSV: {output_path}")
    
    header = ['year', 'month', 'scope', 'activity_name', 'emission_kgco2e', 'emission_tco2e']
    
    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        
        for row in sorted(all_data, key=lambda x: (x['year'], x['month'], x['scope'])):
            writer.writerow([
                row['year'],
                row['month'],
                row['scope'],
                row['activity'],
                f"{row['emission_kgco2e']:.6f}",
                f"{row['emission_tco2e']:.8f}"
            ])
    
    print(f"Written {len(all_data)} rows")

def analyze_excel_detailed(excel_path):
    """Detailed analysis of Excel structure."""
    print("\n" + "="*60)
    print("DETAILED EXCEL STRUCTURE ANALYSIS")
    print("="*60)
    
    wb = load_workbook(excel_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        if 'สรุป' not in sheet_name:
            continue
        
        ws = wb[sheet_name]
        print(f"\n=== Sheet: {sheet_name} ===")
        print(f"Dimensions: {ws.dimensions}")
        
        # Print all rows up to row 60 to find the structure
        print("\nRow content (first 15 columns):")
        for row_idx in range(1, min(70, ws.max_row + 1)):
            row_vals = []
            for col_idx in range(1, 16):
                val = get_cell_value(ws, row_idx, col_idx)
                if val is None:
                    row_vals.append('')
                elif isinstance(val, float):
                    row_vals.append(f"{val:.2f}")
                else:
                    row_vals.append(str(val)[:20])
            print(f"Row {row_idx:2}: {' | '.join(row_vals[:10])}")
    
    wb.close()

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = '/home/rae_admin/joomla-greenoffice/exdata/1.5_GreenhouseGas.xlsx'
    output_csv = os.path.join(script_dir, 'ghg_2567-2568_canonical.csv')
    
    # First, analyze the structure in detail
    analyze_excel_detailed(excel_path)
    
    # Then parse and validate
    all_data = parse_excel_v2(excel_path)
    
    if all_data:
        computed = compute_totals(all_data)
        mismatches = validate_and_report(computed)
        
        if not mismatches:
            generate_canonical_csv(all_data, output_csv)
            print("\n✅ Canonical CSV generated successfully!")
        else:
            print("\n⚠️  Mismatches found. Please review extraction logic.")
            for m in mismatches:
                print(f"  - {m}")
    else:
        print("\n❌ No data extracted from Excel. Need to fix extraction logic.")

if __name__ == '__main__':
    main()